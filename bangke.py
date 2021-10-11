# -*- coding: utf-8 -*-
"""
Created on Sun Oct 10 15:27:39 2021

@author: NAY3HC
"""
import os
from openpyxl import load_workbook
import glob
import pandas as pd

direction  = "C:/Users/NAY3HC/Desktop/core test/test python/BANG KE 2021/"
excelfile = 'C:/Users/NAY3HC/Desktop/core test/test python/consolidate.xlsx'
wbfinal = load_workbook(excelfile)
sfinal = wbfinal['Sheet1']


os.chdir(direction)
ii = 2
for file in glob.glob("*.xlsx"):
    
#    print(file)
    link2direc = direction +file
    wb = load_workbook(link2direc)
    
    xl = pd.ExcelFile(link2direc)
    sheets = xl.sheet_names
    for sheetname in sheets:
        
        s1 = wb[sheetname]
        for testi in range(1,50):
            if s1.cell(testi,1).value == 'STT':
                STTlock = testi +1
                #print(STTlock)
                for testii in range(STTlock,3000):
                    if  str(s1.cell(testii,2).value) != 'None'  :
                        donhanglock = testii   
                 
        for i in range(STTlock,donhanglock+1):
            
            STT = s1.cell(i,1).value
            tenhang = s1.cell(i,3).value
            dvt = s1.cell(i,4).value
            soluong = s1.cell(i,5).value
            dongia = s1.cell(i,6).value
            thanhtien = s1.cell(i,7).value     
            
             
            
              
            sfinal.cell(ii,1).value = 'T ' + file[4:6]
            sfinal.cell(ii,2).value = STT 
            sfinal.cell(ii,3).value = tenhang
            sfinal.cell(ii,4).value = dvt
            sfinal.cell(ii,5).value = soluong
            sfinal.cell(ii,6).value = dongia
            sfinal.cell(ii,7).value = thanhtien
            ii = ii +1
    
    





wbfinal.save(excelfile)    



